from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
sheet = wb.active
sheet.title = "2025 매출 계획"

# 제목
sheet['A1'] = '2025년 월별 매출 계획'
sheet['A1'].font = Font(bold=True, size=14)
sheet.merge_cells('A1:N1')
sheet['A1'].alignment = Alignment(horizontal='center')

# 주요 가정 섹션
sheet['A3'] = '주요 가정'
sheet['A3'].font = Font(bold=True, size=12)

sheet['A4'] = '기준 월 매출 (백만원)'
sheet['B4'] = 100
sheet['B4'].font = Font(color='0000FF')
sheet['B4'].fill = PatternFill('solid', start_color='FFFF00')
sheet['B4'].number_format = '#,##0'

sheet['A5'] = '월별 성장률 (%)'
sheet['B5'] = 5
sheet['B5'].font = Font(color='0000FF')
sheet['B5'].fill = PatternFill('solid', start_color='FFFF00')
sheet['B5'].number_format = '0.0%'

# 월별 헤더
sheet['A7'] = '구분'
months = ['1월', '2월', '3월', '4월', '5월', '6월', '7월', '8월', '9월', '10월', '11월', '12월', '연간 합계']
for i, month in enumerate(months, start=2):
    cell = sheet.cell(row=7, column=i)
    cell.value = month
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 매출 계획 행
sheet['A8'] = '월별 매출'
for col in range(2, 14):
    cell = sheet.cell(row=8, column=col)
    if col == 2:
        cell.value = f'=$B$4'
    else:
        prev_col = get_column_letter(col-1)
        cell.value = f'={prev_col}8*(1+$B$5)'
    cell.font = Font(color='000000')
    cell.number_format = '#,##0'

# 연간 합계
sheet['N8'] = '=SUM(B8:M8)'
sheet['N8'].font = Font(bold=True)
sheet['N8'].number_format = '#,##0'

# 분기별 합계
sheet['A10'] = '분기별 합계'
quarters = [('Q1', 'B8:D8'), ('Q2', 'E8:G8'), ('Q3', 'H8:J8'), ('Q4', 'K8:M8')]
for i, (quarter, range_ref) in enumerate(quarters, start=2):
    sheet.cell(row=10, column=i+(i-2)*2).value = quarter
    sheet.cell(row=10, column=i+(i-2)*2).font = Font(bold=True)
    sheet.cell(row=10, column=i+(i-2)*2+1).value = f'=SUM({range_ref})'
    sheet.cell(row=10, column=i+(i-2)*2+1).number_format = '#,##0'

# 평균 월 매출
sheet['A12'] = '평균 월 매출'
sheet['B12'] = '=AVERAGE(B8:M8)'
sheet['B12'].number_format = '#,##0'

# 최고 매출 월
sheet['A13'] = '최고 매출'
sheet['B13'] = '=MAX(B8:M8)'
sheet['B13'].number_format = '#,##0'

# 최저 매출 월
sheet['A14'] = '최저 매출'
sheet['B14'] = '=MIN(B8:M8)'
sheet['B14'].number_format = '#,##0'

# 열 너비 조정
sheet.column_dimensions['A'].width = 20
for col in range(2, 15):
    sheet.column_dimensions[get_column_letter(col)].width = 12

# 테두리 추가
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in range(7, 15):
    for col in range(1, 15):
        sheet.cell(row=row, column=col).border = thin_border

wb.save('2025_매출계획.xlsx')
print("스프레드시트가 생성되었습니다: 2025_매출계획.xlsx")
