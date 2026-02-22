#!/usr/bin/env python3
"""
HWP 자동입력 매크로 생성기
엑셀 파일에서 변수와 값을 읽어 HWP 매크로 코드가 포함된 .md 파일을 시트별로 생성합니다.

Usage:
    python generate_macros.py <excel_file> [output_dir] [value_column]

Args:
    excel_file:   엑셀 파일 경로
    output_dir:   출력 디렉토리 (기본: 엑셀 파일 옆 output/ 폴더)
    value_column: 값이 있는 열 번호 (기본: 3 = C열)
"""
import sys
import os
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)


def parse_edu_future(text):
    """{{edu_future}} 내용을 [Basic], [Advance], [Expert] 섹션별로 파싱하여
    서브 변수를 추출한다.
    Returns: list of (key, value) tuples"""
    if not text:
        return []

    sections = {}
    current_section = None
    lines_in_section = []

    for line in str(text).split('\n'):
        stripped = line.strip()
        if stripped in ('[Basic]', '[Advance]', '[Expert]'):
            if current_section and lines_in_section:
                sections[current_section] = lines_in_section
            current_section = stripped[1:-1]  # Remove brackets
            lines_in_section = []
        elif stripped and current_section:
            lines_in_section.append(stripped)

    if current_section and lines_in_section:
        sections[current_section] = lines_in_section

    # Map section lines to variable names
    sub_vars = []
    section_map = {
        'Basic':   [('{{edu_future-Basic-1}}', 0), ('{{edu_future-Basic-2}}', 1), ('{{edu_future-Basic-3}}', 2)],
        'Advance': [('{{edu_future-Advance-4}}', 0), ('{{edu_future-Advance-5}}', 1), ('{{edu_future-Advance-6}}', 2), ('{{edu_future-Advance-7}}', 3)],
        'Expert':  [('{{edu_future-Expert-1}}', 0)],
    }

    for section, mappings in section_map.items():
        lines = sections.get(section, [])
        for var_name, idx in mappings:
            value = lines[idx] if idx < len(lines) else ""
            sub_vars.append((var_name, value))

    return sub_vars


def format_value_for_macro(value):
    """값을 HWP 매크로 JS 문자열 형식으로 변환.
    여러 줄이면 \\r 구분자 + 문자열 연결 패턴 적용."""
    if value is None or str(value).strip() == "":
        return '""'

    text = str(value)
    # 큰따옴표 이스케이프
    text = text.replace('"', '\\"')
    lines = text.split('\n')

    if len(lines) == 1:
        return f'"{text}"'

    parts = []
    for i, line in enumerate(lines):
        if i < len(lines) - 1:
            parts.append(f'      "{line}\\r" +')
        else:
            parts.append(f'      "{line}"')
    return '\n'.join(parts)


def generate_macro_code(var_pairs):
    """변수-값 쌍 리스트로부터 HWP 매크로 코드 생성.
    var_pairs: list of (key, value) tuples"""
    map_entries = []
    for key, value in var_pairs:
        formatted = format_value_for_macro(value)
        if '\n' in formatted:
            map_entries.append(f'    "{key}":\n{formatted}')
        else:
            map_entries.append(f'    "{key}": {formatted}')

    map_str = ',\n'.join(map_entries)

    return f'''function OnScriptMacro_자동입력매크로() {{
  // 치환 맵: 값 중 줄바꿈은 \\r 로 표기 (한글에서 문단 구분)
  var map = {{
{map_str}
  }};

  // 토큰을 하나씩 검색 → 삭제 → InsertText로 다줄 텍스트 삽입
  for (var key in map) {{
    while (true) {{
      // 토큰 찾기
      HAction.GetDefault("RepeatFind", HParameterSet.HFindReplace.HSet);
      HParameterSet.HFindReplace.FindString = key;
      HParameterSet.HFindReplace.Direction = 1;       // forward
      HParameterSet.HFindReplace.MatchCase = 0;
      HParameterSet.HFindReplace.WholeWordOnly = 0;
      HParameterSet.HFindReplace.IgnoreMessage = 1;   // 메시지 숨김

      if (!HAction.Execute("RepeatFind", HParameterSet.HFindReplace.HSet)) break;

      // 토큰 삭제
      HAction.Run("Delete");

      // 텍스트 삽입 (여러 줄은 \\r 로 구분)
      HAction.GetDefault("InsertText", HParameterSet.HInsertText.HSet);
      HParameterSet.HInsertText.Text = map[key];
      HAction.Execute("InsertText", HParameterSet.HInsertText.HSet);
    }}
  }}

  return true;
}}'''


def process_excel(excel_path, output_dir, value_col=3, var_start_row=3, var_end_row=41):
    """엑셀 파일을 처리하여 시트별 .md 파일 생성."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    generated = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        var_pairs = []
        edu_future_value = None

        for row in range(var_start_row, var_end_row + 1):
            var_name = ws.cell(row=row, column=1).value
            if var_name and re.match(r'\{\{.+\}\}', str(var_name).strip()):
                key = str(var_name).strip()
                value = ws.cell(row=row, column=value_col).value

                if key == '{{edu_future}}':
                    edu_future_value = value
                    # edu_future 원본은 맵에 넣지 않고, 서브 변수만 추가
                    continue

                var_pairs.append((key, value))

        # {{edu_future}} 서브 변수 추가
        if edu_future_value:
            var_pairs.extend(parse_edu_future(edu_future_value))

        if not var_pairs:
            continue

        macro_code = generate_macro_code(var_pairs)

        safe_name = re.sub(r'[<>:"/\\|?*]', '_', sheet_name.strip())
        md_path = Path(output_dir) / f"{safe_name}.md"

        md_content = f"""# {sheet_name} - HWP 자동입력 매크로

## 사용법
1. 한글(HWP/HWPX) 문서를 엽니다
2. 도구 > 스크립트 매크로 에서 아래 코드를 붙여넣고 실행합니다

## 매크로 코드

```javascript
{macro_code}
```
"""
        md_path.write_text(md_content, encoding='utf-8')
        generated.append(safe_name)
        print(f"  Generated: {md_path.name}")

    return generated


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_macros.py <excel_file> [output_dir] [value_column]")
        print("  excel_file:   엑셀 파일 경로")
        print("  output_dir:   출력 디렉토리 (기본: 엑셀 파일 옆 output/ 폴더)")
        print("  value_column: 값이 있는 열 번호 (기본: 3 = C열)")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(os.path.abspath(excel_file)), "output")
    value_col = int(sys.argv[3]) if len(sys.argv) > 3 else 3

    if not os.path.exists(excel_file):
        print(f"Error: File not found: {excel_file}")
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)
    print(f"Input:  {excel_file}")
    print(f"Output: {output_dir}")
    print(f"Value column: {value_col} ({'ABCDEFGHIJ'[value_col-1]}열)")
    print()

    results = process_excel(excel_file, output_dir, value_col)
    print(f"\nTotal: {len(results)} files generated")


if __name__ == "__main__":
    main()
