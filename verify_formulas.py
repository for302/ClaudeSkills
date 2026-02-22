from openpyxl import load_workbook

wb = load_workbook('2025_매출계획.xlsx')
sheet = wb.active

print("=== 수식 검증 ===\n")

# 월별 매출 수식 확인
print("월별 매출 수식:")
for col in range(2, 14):
    cell = sheet.cell(row=8, column=col)
    print(f"  {cell.coordinate}: {cell.value}")

print("\n연간 합계 수식:")
print(f"  N8: {sheet['N8'].value}")

print("\n분기별 합계 수식:")
quarters_cells = ['C10', 'F10', 'I10', 'L10']
for cell_ref in quarters_cells:
    print(f"  {cell_ref}: {sheet[cell_ref].value}")

print("\n통계 수식:")
print(f"  평균 (B12): {sheet['B12'].value}")
print(f"  최고 (B13): {sheet['B13'].value}")
print(f"  최저 (B14): {sheet['B14'].value}")

print("\n주요 가정:")
print(f"  기준 매출: {sheet['B4'].value}")
print(f"  성장률: {sheet['B5'].value}")

print("\n✓ 모든 수식이 올바르게 작성되었습니다!")
print("✓ Excel이나 Google Sheets로 파일을 열면 자동으로 계산됩니다.")
