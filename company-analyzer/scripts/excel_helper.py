"""
Helper functions for reading and writing company information to Excel files.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


def read_companies(filepath, sheet_name=0):
    """
    Read company information from Excel file with UTF-8 encoding.

    Args:
        filepath: Path to Excel file
        sheet_name: Sheet name or index (default: 0)

    Returns:
        DataFrame with company information

    Expected columns:
        - 기업명 (Company Name)
        - 홈페이지주소 (Homepage URL) - optional
        - 대표자명 (CEO Name) - optional

    Note:
        Uses openpyxl engine with dtype=str to ensure Korean company names
        are read correctly without encoding issues.
    """
    df = pd.read_excel(filepath, sheet_name=sheet_name, engine='openpyxl', dtype=str)

    # Check required columns
    if '기업명' not in df.columns:
        raise ValueError("Excel must contain '기업명' column")

    return df


def write_analysis_results(filepath, results, sheet_name=0):
    """
    Write analysis results to Excel file.

    Args:
        filepath: Path to Excel file
        results: DataFrame with analysis results
        sheet_name: Sheet name or index (default: 0)

    Expected columns in results DataFrame:
        - 기업명 (Company Name)
        - 사업분야및주요제품서비스 (Business Areas and Main Products/Services)
        - 주요고객및시장현황 (Main Customers and Market Status)
        - 국내시장현황 (Domestic Market Status)
        - 글로벌시장현황 (Global Market Status)
        - 기술역량 (Technology Capability)
        - 수상실적 (Awards and Recognition)
        - 뉴스요약 (News and Updates Summary)
    """
    # Load existing workbook
    wb = load_workbook(filepath)

    if isinstance(sheet_name, str):
        sheet = wb[sheet_name]
    else:
        sheet = wb.worksheets[sheet_name]

    # Write results to Excel
    results.to_excel(filepath, sheet_name=sheet_name, index=False)

    # Reload to apply formatting
    wb = load_workbook(filepath)
    sheet = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    # Format headers
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Enable text wrapping for analysis columns
    analysis_columns = ['사업분야및주요제품서비스', '주요고객및시장현황', '국내시장현황',
                       '글로벌시장현황', '기술역량', '수상실적', '뉴스요약']

    for col_name in analysis_columns:
        if col_name in results.columns:
            col_idx = results.columns.get_loc(col_name) + 1
            for row in range(2, len(results) + 2):
                cell = sheet.cell(row=row, column=col_idx)
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    wb.save(filepath)
    print(f"Results written to {filepath}")


def get_companies_to_analyze(filepath, sheet_name=0):
    """
    Get list of companies that need analysis.

    Args:
        filepath: Path to Excel file
        sheet_name: Sheet name or index (default: 0)

    Returns:
        List of dictionaries with company information
    """
    df = read_companies(filepath, sheet_name)

    companies = []
    for _, row in df.iterrows():
        company = {
            '기업명': row.get('기업명', ''),
            '홈페이지주소': row.get('홈페이지주소', ''),
            '대표자명': row.get('대표자명', '')
        }
        companies.append(company)

    return companies


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_helper.py <excel_file>")
        sys.exit(1)

    filepath = sys.argv[1]

    try:
        companies = get_companies_to_analyze(filepath)
        print(f"Found {len(companies)} companies to analyze:")
        for i, company in enumerate(companies, 1):
            print(f"{i}. {company['기업명']}")
            if company['홈페이지주소']:
                print(f"   Homepage: {company['홈페이지주소']}")
            if company['대표자명']:
                print(f"   CEO: {company['대표자명']}")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
